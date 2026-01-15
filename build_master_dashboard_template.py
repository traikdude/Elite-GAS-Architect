from __future__ import annotations
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = "Master_Automation_Dashboard_Template.xlsx"

SHEETS = [
    "README",
    "Dashboard",
    "Config",
    "Action_Queue",
    "Master Dashboard",
    "Enhancement Reports",
]

MASTER_DASHBOARD_HEADERS = [
    "timestamp_iso_ms",
    "date_local",
    "time_local",
    "epoch_ms",
    "event_type",
    "action",
    "user",
    "status",
    "duration_ms",
    "remaining_quota_hint",
    "error_message",
    "stack_trace",
    "meta_json",
]

ENHANCEMENT_REPORT_HEADERS = [
    "created_iso_ms",
    "created_by",
    "work_product_title",
    "source",
    "word_count",
    "signals_json",
    "analysis_markdown",
    "proposal_markdown",
    "prompt_markdown",
    "ai_response_markdown",
]

ACTION_QUEUE_HEADERS = [
    "queue_id",
    "created_iso_ms",
    "requested_by",
    "action_type",
    "target_type",
    "target_id",
    "params_json",
    "priority",
    "status",
    "status_message",
    "result_link",
    "result_id",
    "started_iso_ms",
    "completed_iso_ms",
    "duration_ms",
    "correlation_id",
]

def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="111827")
    font = Font(bold=True, color="F9FAFB")
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = ws["A2"]

def autosize(ws, min_w=10, max_w=70):
    for col in range(1, ws.max_column + 1):
        maxlen = 0
        for row in range(1, min(ws.max_row, 200) + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            maxlen = max(maxlen, len(str(v)))
        width = max(min_w, min(max_w, maxlen + 2))
        ws.column_dimensions[get_column_letter(col)].width = width

def add_named_list(ws, name, values, start_row, col):
    # Writes values vertically and returns the A1 range string.
    for i, v in enumerate(values):
        ws.cell(row=start_row + i, column=col, value=v)
    a1_start = f"{get_column_letter(col)}{start_row}"
    a1_end   = f"{get_column_letter(col)}{start_row + len(values) - 1}"
    return f"'{ws.title}'!{a1_start}:{a1_end}"

def create():
    wb = Workbook()
    # Remove default
    wb.remove(wb.active)

    for s in SHEETS:
        wb.create_sheet(s)

    # ===== README =====
    r = wb["README"]
    r["A1"] = "Master Automation — Sheets-Native Dashboard Template"
    r["A1"].font = Font(bold=True, size=14)
    r["A3"] = "How to use"
    r["A3"].font = Font(bold=True, size=12)
    r["A4"] = "1) Use the Dashboard sheet to select an ActionType, TargetType, TargetId, and Params JSON."
    r["A5"] = "2) Set Enqueue = TRUE to request execution (your Apps Script onEdit handler will append to Action_Queue)."
    r["A6"] = "3) A separate processor (time trigger / menu action) should pull queued rows and run actions."
    r["A7"] = "4) Log all activity to 'Master Dashboard' and write any outputs/results back to the queue row."
    r["A9"] = "Core design goal: the spreadsheet is the control plane; Apps Script is the execution engine."
    r["A11"] = "Sheets in this workbook"
    r["A11"].font = Font(bold=True)
    for i, s in enumerate(SHEETS, start=12):
        r[f"A{i}"] = f"- {s}"
    autosize(r)

    # ===== CONFIG =====
    c = wb["Config"]
    c["A1"] = "Settings"
    c["A1"].font = Font(bold=True, size=12)
    c["A2"] = "key"; c["B2"] = "value"
    style_header(c, row=2)

    settings = [
        ("system_name", "Master Automation Suite"),
        ("timezone", "America/New_York"),
        ("default_priority", "NORMAL"),
        ("queue_processing_mode", "MANUAL"),
        ("max_batch_size", "25"),
    ]
    for idx, (k, v) in enumerate(settings, start=3):
        c[f"A{idx}"] = k
        c[f"B{idx}"] = v

    # Action Catalog
    start = 10
    c[f"A{start}"] = "Action Catalog"
    c[f"A{start}"].font = Font(bold=True, size=12)
    headers = ["action_type","description","handler_function","default_target_type","params_schema_json","enabled","requires_auth_scopes","notes"]
    for j, h in enumerate(headers, start=1):
        c.cell(row=start+1, column=j, value=h)
    style_header(c, row=start+1)

    catalog = [
        ("OPEN_LINK","Open a configured link key or URL","MASTER_apiOpenLinkKey","LINK_KEY",'{"key":"COLAB|GITHUB|WEBAPP"}', True, "script.container.ui", "Uses modeless dialog opener"),
        ("CREATE_PROJECT_FOLDER","Create Drive folder structure","MASTER_apiCreateProjectFolder","SPREADSHEET",'{}', True, "drive", "Stores folder id in Document Properties"),
        ("GENERATE_ENHANCEMENT","Generate enhancement package from text","MASTER_apiGenerateEnhancement","TEXT",'{"title":"","source":"","callAi":false}', True, "script.external_request", "Optional AI endpoint call"),
        ("OPEN_DASHBOARD","Activate Master Dashboard sheet","MASTER_openDashboard","SHEET",'{}', True, "spreadsheets.currentonly", "Navigation helper"),
    ]
    for i, row in enumerate(catalog, start=start+2):
        for j, v in enumerate(row, start=1):
            c.cell(row=i, column=j, value=v)

    # Lists for validation
    list_col = 12  # Column L
    c["L2"] = "Lists (for data validation)"
    c["L2"].font = Font(bold=True)
    action_types = [r[0] for r in catalog] + ["CUSTOM"]
    target_types = ["SPREADSHEET","SHEET","RANGE","DOC","GMAIL","DRIVE","SITE","LINK_KEY","TEXT","FILE"]
    priorities   = ["LOW","NORMAL","HIGH","URGENT"]
    statuses     = ["NEW","QUEUED","RUNNING","DONE","ERROR","CANCELLED"]

    at_rng = add_named_list(c, "ActionTypes", action_types, start_row=4, col=list_col)
    tt_rng = add_named_list(c, "TargetTypes", target_types, start_row=4, col=list_col+1)
    pr_rng = add_named_list(c, "Priorities", priorities,   start_row=4, col=list_col+2)
    st_rng = add_named_list(c, "Statuses", statuses,       start_row=4, col=list_col+3)

    # Name the ranges
    wb.create_named_range("ActionTypes", c, at_rng.split("!")[1])
    wb.create_named_range("TargetTypes", c, tt_rng.split("!")[1])
    wb.create_named_range("Priorities",  c, pr_rng.split("!")[1])
    wb.create_named_range("Statuses",    c, st_rng.split("!")[1])

    autosize(c)

    # ===== DASHBOARD =====
    d = wb["Dashboard"]
    d["A1"] = "Dashboard Control Panel"
    d["A1"].font = Font(bold=True, size=14)

    labels = [
        ("A3","Request ID (auto)"),
        ("A4","Action Type"),
        ("A5","Target Type"),
        ("A6","Target ID / Key"),
        ("A7","Params (JSON)"),
        ("A8","Priority"),
        ("A9","Enqueue (TRUE)"),
        ("A11","Status"),
        ("A12","Last Message"),
        ("A13","Result Link"),
    ]
    for addr, text in labels:
        d[addr] = text
        d[addr].font = Font(bold=True)
        d[addr].alignment = Alignment(vertical="center")

    input_fill = PatternFill("solid", fgColor="DBEAFE")  # light blue
    for addr in ["B4","B5","B6","B7","B8","B9"]:
        d[addr].fill = input_fill

    # default values
    d["B3"] = "=TEXT(NOW(),\"yyyymmdd-hhmmss\")"
    d["B8"] = "NORMAL"
    d["B9"] = "FALSE"

    # output placeholders
    d["B11"] = "=IF(B9=TRUE,\"QUEUED\",\"(not queued)\")"
    d["B12"] = "=IF(B9=TRUE,\"Waiting for processor…\",\"—\")"
    d["B13"] = ""

    # Data validations (use named ranges)
    dv_action = DataValidation(type="list", formula1="=ActionTypes", allow_blank=False)
    dv_target = DataValidation(type="list", formula1="=TargetTypes", allow_blank=False)
    dv_prio   = DataValidation(type="list", formula1="=Priorities",  allow_blank=False)
    dv_bool   = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=False)

    d.add_data_validation(dv_action); dv_action.add(d["B4"])
    d.add_data_validation(dv_target); dv_target.add(d["B5"])
    d.add_data_validation(dv_prio);   dv_prio.add(d["B8"])
    d.add_data_validation(dv_bool);   dv_bool.add(d["B9"])

    d.column_dimensions["A"].width = 22
    d.column_dimensions["B"].width = 70
    d.row_dimensions[7].height = 60
    d["B7"].alignment = Alignment(wrap_text=True, vertical="top")
    d["B7"].font = Font(name="Consolas")
    autosize(d)

    # ===== ACTION QUEUE =====
    q = wb["Action_Queue"]
    for j, h in enumerate(ACTION_QUEUE_HEADERS, start=1):
        q.cell(row=1, column=j, value=h)
    style_header(q, row=1)

    # Example rows
    example = [
        ["Q-0001", "=TEXT(NOW(),\"yyyy-mm-dd\\\"T\\\"hh:mm:ss.000Z\")", "you@domain.com", "OPEN_LINK", "LINK_KEY", "WEBAPP", '{"key":"WEBAPP"}', "NORMAL", "QUEUED", "", "", "", "", "", "", "C-0001"],
        ["Q-0002", "=TEXT(NOW(),\"yyyy-mm-dd\\\"T\\\"hh:mm:ss.000Z\")", "you@domain.com", "CREATE_PROJECT_FOLDER", "SPREADSHEET", "", "{}", "NORMAL", "NEW", "", "", "", "", "", "", "C-0002"],
    ]
    for i, row in enumerate(example, start=2):
        for j, v in enumerate(row, start=1):
            q.cell(row=i, column=j, value=v)

    # validations for queue
    dv_q_action = DataValidation(type="list", formula1="=ActionTypes", allow_blank=False)
    dv_q_target = DataValidation(type="list", formula1="=TargetTypes", allow_blank=False)
    dv_q_prio   = DataValidation(type="list", formula1="=Priorities",  allow_blank=False)
    dv_q_status = DataValidation(type="list", formula1="=Statuses",    allow_blank=False)
    q.add_data_validation(dv_q_action); dv_q_action.add(f"D2:D5000")
    q.add_data_validation(dv_q_target); dv_q_target.add(f"E2:E5000")
    q.add_data_validation(dv_q_prio);   dv_q_prio.add(f"H2:H5000")
    q.add_data_validation(dv_q_status); dv_q_status.add(f"I2:I5000")

    autosize(q)

    # ===== MASTER DASHBOARD (LOGS) =====
    m = wb["Master Dashboard"]
    for j, h in enumerate(MASTER_DASHBOARD_HEADERS, start=1):
        m.cell(row=1, column=j, value=h)
    style_header(m, row=1)
    autosize(m)

    # ===== ENHANCEMENT REPORTS =====
    e = wb["Enhancement Reports"]
    for j, h in enumerate(ENHANCEMENT_REPORT_HEADERS, start=1):
        e.cell(row=1, column=j, value=h)
    style_header(e, row=1)
    autosize(e)

    wb.save(OUT)
    print(f"✅ Wrote: {OUT}")

if __name__ == "__main__":
    create()
